#!/usr/bin/env python3
# SAP-Role-Updater - Combined AGR_1251/AGR_1252 modifier (fixed-width SAP role exports)
# - Single rules workbook can target AGR_1251 and/or AGR_1252 in one run.
# - Preserves 1:1 every line that is not the targeted table.
# - Only action supported: replace_list.
# - Log CSV uses ';' delimiter with header: action;before;after.
# Rules columns (case-insensitive): ACTION, TABLE, MANDT, AGR_NAME, OBJECT, AUTH, FIELD, LOW, HIGH
#   * AGR_1251: OBJECT/AUTH required; FIELD = auth field; LOW/HIGH = value or range (40 chars each).
#   * AGR_1252: OBJECT/AUTH empty; FIELD = org field (e.g. $WERKS); LOW/HIGH = org ranges/values.
# Usage example:
#   python SAP-Role-Updater.py --in EXPORT.txt --rules RULES.xlsx --outdir ./salida

import csv
import json
import os
import re
from collections import defaultdict
from collections.abc import Callable
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook

from sap_role_updater.core.constants import PREFIX_WIDTHS, RX_1251, RX_1252, RX_1252_LEGACY, W1251, W1252
from sap_role_updater.gui.i18n import t
from sap_role_updater.utils.error_handler import CodedError, raise_error
from sap_role_updater.utils.path_safety import (
    DEFAULT_LIMITS,
    resolve_output_dir,
    resolve_regular_file,
    safe_output_path,
    sha256_file,
)
from sap_role_updater.version import APP_VERSION

__version__ = APP_VERSION

def read_text(path):
    """Read text tolerating common encodings; strip CR/LF endings."""
    for enc in ("utf-8", "latin-1", "cp1252"):
        try:
            with open(path, encoding=enc, newline="") as f:
                return [ln.rstrip("\r\n") for ln in f.readlines()], enc
        except UnicodeDecodeError:
            continue
    with open(path, "rb") as f:
        data = f.read()
    return data.decode("latin-1", errors="ignore").splitlines(), "latin-1"


def fmt_fixed(val: str, width: int) -> str:
    s = val or ""
    return s[:width] if len(s) > width else s.ljust(width)


def split_sequence(raw_vals: str, keep_empty: bool = False):
    """Split on | or , preserving item positions when requested.

    keep_empty=True is used for HIGH to allow mixed rules such as:
      LOW  = *|/*|0*|A*
      HIGH = ||9*|Z*
    """
    txt = raw_vals or ""
    if not txt.strip() and "|" not in txt and "," not in txt:
        return []

    if "|" in txt:
        parts = txt.split("|")
    elif "," in txt:
        parts = txt.split(",")
    else:
        parts = [txt]

    vals = [p.strip() for p in parts]
    if keep_empty:
        return vals
    return [v for v in vals if v != ""]


def split_pairs(raw_low: str, raw_high: str, rule_ctx: dict = None):
    """Pair LOW/HIGH lists; HIGH may be shorter/empty, defaults to ''.
    If both are empty, return a single empty pair to force replace with blanks.
    If HIGH is provided without LOW, raise a clear validation error."""
    lows = split_sequence(raw_low, keep_empty=False)
    highs = split_sequence(raw_high, keep_empty=True)
    if not lows and any(h != "" for h in highs):
        ctx = rule_ctx or {}
        raise_error(
            "VAL-002",
            "SEV2",
            t("val.high_without_low"),
            err_type="Validation",
            origin="split_pairs",
            details=(
                f"row={ctx.get('row', '?')}, "
                f"table={ctx.get('table', '?')}, "
                f"role={ctx.get('role', '?')}, "
                f"field={ctx.get('field', '?')}"
            ),
        )
    pairs = []
    if not lows and not highs:
        pairs.append(("", ""))
    else:
        for idx, low_val in enumerate(lows):
            high_val = highs[idx] if idx < len(highs) else ""
            pairs.append((low_val, high_val))
    # de-dupe by (low, high) preserving order
    seen = set()
    dedup = []
    for low_val, high_val in pairs:
        key = (low_val, high_val)
        if key not in seen:
            seen.add(key)
            dedup.append(key)
    return dedup


def append_replace_logs(befores, afters, log_rows):
    """Align before/after lists and log REPLACE rows duplicating the shorter side."""
    if not befores and not afters:
        return
    n = max(len(befores), len(afters))
    for i in range(n):
        b = befores[min(i, len(befores) - 1)] if befores else ""
        a = afters[min(i, len(afters) - 1)] if afters else ""
        log_rows.append(["REPLACE", b, a])


def build_output_paths(infile, outdir):
    base = os.path.basename(infile)
    name, ext = os.path.splitext(base)
    if not ext:
        ext = ""
    outfile = os.path.join(outdir, f"{name}_MOD{ext}")
    log_path = os.path.join(outdir, f"{name}_MOD_LOG.csv")
    return outfile, log_path


def build_meta_path(infile, outdir):
    base = os.path.basename(infile)
    name, _ = os.path.splitext(base)
    return os.path.join(outdir, f"{name}_MOD_META.json")


def _empty_result(status="ok"):
    return {
        "status": status,
        "counters": {"adds": 0, "deletes": 0, "replaces": 0, "warns": 0},
        "outfile": "",
        "log_path": "",
        "has_validation_errors": False,
        "warns_details": [],
        "warns_struct": [],
        "sample_rows": [],
        "encoding_detected": "",
        "rules_sheet_detected": "",
        "base_stats": {},
        "rules_stats": {},
        "coverage_rows": [],
        "coverage_summary": {},
        "checksums": {},
        "meta_path": "",
        "path_warnings": [],
        "privacy_mode": False,
    }


def _build_base_stats(lines, entries):
    roles = set()
    cnt_1251 = 0
    cnt_1252 = 0
    for e in entries:
        if not e:
            continue
        role = e.get("role", "").strip()
        if role:
            roles.add(role)
        if e.get("table_type") == "AGR_1251":
            cnt_1251 += 1
        elif e.get("table_type") == "AGR_1252":
            cnt_1252 += 1
    return {
        "total_lines": len(lines),
        "target_lines": cnt_1251 + cnt_1252,
        "agr_1251_lines": cnt_1251,
        "agr_1252_lines": cnt_1252,
        "other_lines": len(lines) - (cnt_1251 + cnt_1252),
        "roles_unique": len(roles),
    }


def _coverage_status_counts(coverage_rows):
    counts = {
        "total_rules": len(coverage_rows),
        "applied": 0,
        "no_base": 0,
        "skipped_error": 0,
        "cancelled": 0,
    }
    for row in coverage_rows:
        status = row.get("status", "")
        if status == "APPLIED":
            counts["applied"] += 1
        elif status == "NO_BASE":
            counts["no_base"] += 1
        elif status == "SKIPPED_ERROR":
            counts["skipped_error"] += 1
        elif status == "CANCELLED":
            counts["cancelled"] += 1
    return counts


def _redact_token(token: str) -> str:
    clean = (token or "").strip()
    if not clean:
        return ""
    if len(clean) <= 3:
        return "*" * len(clean)
    return clean[:3] + "***"


def _redact_line(line: str) -> str:
    entry_1251 = parse_entry_1251(line)
    if entry_1251:
        return compose_line_1251(
            entry_1251,
            entry_1251["counter"],
            entry_1251["field"].strip(),
            _redact_token(entry_1251["low"]),
            _redact_token(entry_1251["high"]),
        )
    entry_1252 = parse_entry_1252(line)
    if entry_1252:
        return compose_line_1252(
            entry_1252,
            entry_1252["counter"],
            entry_1252["varbl"].strip(),
            _redact_token(entry_1252["low"]),
            _redact_token(entry_1252["high"]),
        )
    return line


def _build_metadata_payload(res, *, infile_name, rules_name, outfile_name, log_name):
    return {
        "app_version": APP_VERSION,
        "generated_at_utc": datetime.now(UTC).isoformat(),
        "privacy_mode": bool(res.get("privacy_mode")),
        "checksums": res.get("checksums", {}),
        "inputs": {
            "base_file": infile_name,
            "rules_file": rules_name,
        },
        "outputs": {
            "mod_file": outfile_name,
            "log_file": log_name,
        },
        "counters": res.get("counters", {}),
        "encoding_detected": res.get("encoding_detected", ""),
        "rules_sheet_detected": res.get("rules_sheet_detected", ""),
        "base_stats": res.get("base_stats", {}),
        "rules_stats": res.get("rules_stats", {}),
        "coverage_summary": res.get("coverage_summary", {}),
    }


def run_job_ex(
    infile: str,
    rules_path: str,
    outdir: str | None = None,
    *,
    preview: bool = False,
    verbose: bool = False,
    ui_sample_limit: int = 300,
    progress_cb: Callable[[int, int, str], None] | None = None,
    is_cancelled: Callable[[], bool] | None = None,
    redact_log: bool = False,
    write_meta: bool = False,
    max_base_file_size_mb: int = DEFAULT_LIMITS["base_size_mb"],
    max_rules_file_size_mb: int = DEFAULT_LIMITS["rules_size_mb"],
    max_base_lines: int = DEFAULT_LIMITS["base_lines"],
    max_rules_lines: int = DEFAULT_LIMITS["rules_lines"],
) -> dict:
    """Validate and optionally process one base export plus one rules file."""
    del verbose  # reserved for future hooks; legacy signature compatibility.
    res = _empty_result(status="ok")
    res["privacy_mode"] = bool(redact_log)
    log_fh = None
    log_writer = None
    temp_log_path = ""
    temp_out_path = ""
    temp_meta_path = ""
    final_out_path = ""
    final_log_path = ""
    final_meta_path = ""
    infile_info = {}
    rules_info = {}
    outdir_info = {}

    def progress(current, total, message):
        if progress_cb:
            progress_cb(current, total, message)

    def clean_text(val):
        return (val or "").replace("\r", " ").replace("\n", " ")

    def log_emit(action, before, after):
        safe_before = _redact_line(before) if redact_log and before else before
        safe_after = _redact_line(after) if redact_log and after else after
        row = [action, clean_text(safe_before), clean_text(safe_after)]
        if len(res["sample_rows"]) < ui_sample_limit:
            res["sample_rows"].append(row)
        if log_writer:
            log_writer.writerow(row)

    def warn_emit(code, detail, severity="SEV3", rule=None, legacy=False, msg_id=None, msg_params=None):
        payload_params = msg_params or {}
        final_detail = detail or (t(msg_id, **payload_params) if msg_id else "")
        info = {
            "code": code,
            "severity": severity,
            "row": (rule or {}).get("row"),
            "table": (rule or {}).get("table", ""),
            "role": (rule or {}).get("role", ""),
            "field": (rule or {}).get("field", ""),
            "detail": final_detail,
            "msg_id": msg_id or "",
            "msg_params": payload_params,
        }
        res["warns_struct"].append(info)
        if legacy:
            res["warns_details"].append(final_detail)

    try:
        if not preview and not outdir:
            raise_error("VAL-004", "SEV2", t("val.outdir_required"), origin="run_job_ex", err_type="Validation")

        progress(0, 6, t("progress.validate_paths"))
        infile_info = resolve_regular_file(
            infile,
            label=t("sec.label.base"),
            max_size_mb=max_base_file_size_mb,
            max_lines=max_base_lines,
        )
        rules_info = resolve_regular_file(
            rules_path,
            label=t("sec.label.rules"),
            max_size_mb=max_rules_file_size_mb,
            max_lines=max_rules_lines,
        )
        infile = str(infile_info["path"])
        rules_path = str(rules_info["path"])
        if not preview:
            outdir_info = resolve_output_dir(outdir, label=t("sec.label.output"))
            outdir = str(outdir_info["path"])

        for path_info, label_key in (
            (infile_info, "sec.label.base"),
            (rules_info, "sec.label.rules"),
            (outdir_info, "sec.label.output"),
        ):
            if path_info and path_info.get("is_unc"):
                warn_emit(
                    "WARN-NETPATH",
                    "",
                    severity="SEV3",
                    legacy=True,
                    msg_id="sec.warn_unc_path",
                    msg_params={"label": t(label_key), "name": path_info.get("name", "")},
                )
                res["counters"]["warns"] += 1
                res["path_warnings"].append(
                    {
                        "code": "WARN-NETPATH",
                        "label": t(label_key),
                        "name": path_info.get("name", ""),
                    }
                )
                log_emit("WARN-NETPATH", t("sec.warn_unc_path", label=t(label_key), name=path_info.get("name", "")), "")

        if write_meta:
            progress(0, 6, t("progress.hash_inputs"))
            res["checksums"] = {
                "base_sha256": sha256_file(infile),
                "rules_sha256": sha256_file(rules_path),
            }

        progress(0, 5, t("progress.read_base"))
        lines, enc = read_text(infile)
        res["encoding_detected"] = enc
        entries = build_entries(lines)
        indexes = build_entry_indexes(entries)
        res["base_stats"] = _build_base_stats(lines, entries)
        res["base_stats"]["index_keys_agr_1251"] = indexes["stats"]["agr_1251_keys"]
        res["base_stats"]["index_keys_agr_1252"] = indexes["stats"]["agr_1252_keys"]

        progress(1, 5, t("progress.parse_rules"))
        rules, rules_meta = parse_rules(rules_path, return_meta=True)
        res["rules_sheet_detected"] = rules_meta.get("rules_sheet_detected", "")
        res["rules_stats"] = rules_meta["rules_stats"]
        res["has_validation_errors"] = bool(rules_meta.get("has_validation_errors", False))
        res["coverage_rows"] = [dict(row) for row in rules_meta.get("coverage_rows", [])]
        coverage_by_row = {row["row"]: row for row in res["coverage_rows"]}
        counters_used = build_counters_state(entries)

        for issue in rules_meta.get("validation_issues", []):
            res["counters"]["warns"] += 1
            warn_emit(
                issue.get("code", "VAL-INVALID-TABLE"),
                issue.get("detail", ""),
                severity=issue.get("severity", "SEV2"),
                rule=issue,
                legacy=True,
                msg_id=issue.get("msg_id", ""),
                msg_params=issue.get("msg_params", {}),
            )
            log_emit(issue.get("code", "VAL-INVALID-TABLE"), issue.get("detail", ""), "")

        res["coverage_summary"] = _coverage_status_counts(res["coverage_rows"])
        if res["has_validation_errors"] and not preview:
            res["status"] = "error"
            res["error"] = CodedError(
                "VAL-RULES-INVALID",
                "SEV2",
                t("val.rules_invalid_process"),
                details=t("val.rules_invalid_process_details"),
                err_type="Validation",
                origin="run_job_ex",
            )
            return res

        if not preview:
            final_out_path, final_log_path = build_output_paths(infile, outdir)
            final_meta_path = build_meta_path(infile, outdir) if write_meta else ""
            final_out_path = str(safe_output_path(outdir, os.path.basename(final_out_path), label=t("sec.label.mod")))
            final_log_path = str(safe_output_path(outdir, os.path.basename(final_log_path), label=t("sec.label.log")))
            if final_meta_path:
                final_meta_path = str(
                    safe_output_path(outdir, os.path.basename(final_meta_path), label=t("sec.label.meta"))
                )
            temp_out_path = final_out_path + ".tmp"
            temp_log_path = final_log_path + ".tmp"
            temp_meta_path = final_meta_path + ".tmp" if final_meta_path else ""
            log_fh = open(temp_log_path, "w", encoding="utf-8", newline="")
            log_writer = csv.writer(log_fh, delimiter=";")
            log_writer.writerow(["action", "before", "after"])

        total_rules = max(len(rules), 1)
        for idx, r in enumerate(rules, start=1):
            if is_cancelled and is_cancelled():
                for coverage_row in res["coverage_rows"]:
                    if coverage_row.get("status") == "PENDING":
                        coverage_row["status"] = "CANCELLED"
                        coverage_row["reason_code"] = "USR-001"
                        coverage_row["reason_msg_id"] = "user.cancelled"
                        coverage_row["reason_params"] = {}
                res["coverage_summary"] = _coverage_status_counts(res["coverage_rows"])
                res["status"] = "cancelled"
                return res
            progress(idx, total_rules, t("progress.process_rule", current=idx, total=total_rules))

            rule_log_rows = []
            coverage_row = coverage_by_row.get(r["row"])
            if r["action"] != "replace_list":
                res["counters"]["warns"] += 1
                detail = t("warn.unsupported_action", action=r["action"], row=r["row"])
                if coverage_row is not None:
                    coverage_row["status"] = "SKIPPED_ERROR"
                    coverage_row["reason_code"] = "WARN-ACTION"
                    coverage_row["reason_msg_id"] = "warn.unsupported_action"
                    coverage_row["reason_params"] = {"action": r["action"], "row": r["row"]}
                warn_emit(
                    "WARN-ACTION",
                    detail,
                    severity="SEV3",
                    rule=r,
                    legacy=True,
                    msg_id="warn.unsupported_action",
                    msg_params={"action": r["action"], "row": r["row"]},
                )
                rule_log_rows.append(["WARN-ACTION", detail, ""])
            elif r["table"] == "AGR_1251":
                handle_rule_1251(
                    r,
                    entries,
                    indexes,
                    counters_used,
                    rule_log_rows,
                    res["counters"],
                    warn_emit,
                    coverage_row,
                )
            elif r["table"] == "AGR_1252":
                handle_rule_1252(
                    r,
                    entries,
                    indexes,
                    counters_used,
                    rule_log_rows,
                    res["counters"],
                    warn_emit,
                    coverage_row,
                )
            else:
                res["counters"]["warns"] += 1
                res["has_validation_errors"] = True
                detail = t("val.row_invalid_table", row=r.get("row", ""), value=r.get("table", ""))
                if coverage_row is not None:
                    coverage_row["status"] = "SKIPPED_ERROR"
                    coverage_row["reason_code"] = "VAL-INVALID-TABLE"
                    coverage_row["reason_msg_id"] = "val.row_invalid_table"
                    coverage_row["reason_params"] = {"row": r.get("row", ""), "value": r.get("table", "")}
                warn_emit(
                    "VAL-INVALID-TABLE",
                    detail,
                    severity="SEV2",
                    rule=r,
                    legacy=True,
                    msg_id="val.row_invalid_table",
                    msg_params={"row": r.get("row", ""), "value": r.get("table", "")},
                )
                rule_log_rows.append(["VAL-INVALID-TABLE", detail, ""])

            for action, before, after in rule_log_rows:
                log_emit(action, before, after)

        if is_cancelled and is_cancelled():
            for coverage_row in res["coverage_rows"]:
                if coverage_row.get("status") == "PENDING":
                    coverage_row["status"] = "CANCELLED"
                    coverage_row["reason_code"] = "USR-001"
                    coverage_row["reason_msg_id"] = "user.cancelled"
                    coverage_row["reason_params"] = {}
            res["coverage_summary"] = _coverage_status_counts(res["coverage_rows"])
            res["status"] = "cancelled"
            return res

        out_lines = []
        for i, e in enumerate(entries):
            if e is None:
                out_lines.append(lines[i])  # original non-target line
            elif not e.get("marked_deleted"):
                out_lines.append(e["raw"])

        if not preview:
            progress(total_rules, total_rules, t("progress.write_mod"))
            with open(temp_out_path, "w", encoding=enc, newline="\n") as f:
                for ln in out_lines:
                    f.write(ln + "\n")

            if log_fh:
                log_fh.flush()
                log_fh.close()
                log_fh = None
            if final_meta_path:
                meta_payload = _build_metadata_payload(
                    res,
                    infile_name=infile_info.get("name", os.path.basename(infile)),
                    rules_name=rules_info.get("name", os.path.basename(rules_path)),
                    outfile_name=os.path.basename(final_out_path),
                    log_name=os.path.basename(final_log_path),
                )
                with open(temp_meta_path, "w", encoding="utf-8", newline="\n") as meta_fh:
                    json.dump(meta_payload, meta_fh, ensure_ascii=False, indent=2)
                    meta_fh.write("\n")
            os.replace(temp_out_path, final_out_path)
            os.replace(temp_log_path, final_log_path)
            if final_meta_path:
                os.replace(temp_meta_path, final_meta_path)
            res["outfile"] = final_out_path
            res["log_path"] = final_log_path
            res["meta_path"] = final_meta_path

        res["coverage_summary"] = _coverage_status_counts(res["coverage_rows"])
        progress(total_rules, total_rules, t("progress.done"))
        return res
    except CodedError as ce:
        res["status"] = "error"
        res["error"] = ce
        return res
    except Exception as ex:  # noqa: BLE001
        wrapped = CodedError(
            "SYS-500",
            "SEV1",
            t("sys.unhandled"),
            details=str(ex),
            err_type="System",
            origin="run_job_ex",
        )
        res["status"] = "error"
        res["error"] = wrapped
        return res
    finally:
        if log_fh:
            log_fh.close()
        if res["status"] != "ok":
            for p in (temp_out_path, temp_log_path, temp_meta_path):
                if p and os.path.exists(p):
                    try:
                        os.remove(p)
                    except OSError:
                        pass


def run_job(
    infile,
    rules_path,
    outdir=None,
    verbose=False,
    preview=False,
    *,
    redact_log=False,
    write_meta=False,
):
    res = run_job_ex(
        infile=infile,
        rules_path=rules_path,
        outdir=outdir,
        preview=preview,
        verbose=verbose,
        redact_log=redact_log,
        write_meta=write_meta,
    )
    if res["status"] == "cancelled":
        raise_error("USR-001", "SEV3", t("user.cancelled"), origin="run_job", err_type="User")
    if res["status"] == "error":
        err = res.get("error")
        if isinstance(err, CodedError):
            raise err
        raise_error("SYS-500", "SEV1", t("sys.unhandled"), details=str(err), origin="run_job", err_type="System")
    return res["counters"], res["outfile"], res["log_path"], res["warns_details"]


# ---------------- parse entries ----------------

def parse_entry_1251(line):
    m = RX_1251.match(line)
    if not m:
        return None
    return {
        "table_type": "AGR_1251",
        "raw": line,
        "table": m.group("table"),
        "sp40": m.group("sp40"),
        "mandt": m.group("mandt"),
        "role": m.group("role"),
        "counter": m.group("counter"),
        "object": m.group("object"),
        "auth": m.group("auth"),
        "variant_pad": m.group("variant_pad"),
        "field": m.group("field"),
        "low": m.group("low"),
        "high": m.group("high"),
        "modified": m.group("modified"),
        "deleted": m.group("deleted"),
        "copied": m.group("copied"),
        "neu": m.group("neu"),
        "node": m.group("node"),
        "tail": m.group("tail"),
        "marked_deleted": False,
    }


def parse_entry_1252(line):
    m = RX_1252.match(line)
    if not m:
        # Try legacy without HIGH and 4-char LOW
        m2 = RX_1252_LEGACY.match(line)
        if not m2:
            return None
        m = m2
        high_val = ""
    else:
        high_val = m.group("high")
    if m.group("table").strip() != "AGR_1252":
        return None
    return {
        "table_type": "AGR_1252",
        "raw": line,
        "table": m.group("table"),
        "sp40": m.group("sp40"),
        "mandt": m.group("mandt"),
        "role": m.group("role"),
        "counter": m.group("counter"),
        "varbl": m.group("varbl"),
        "sp30": m.group("sp30"),
        "low": m.group("low"),
        "high": high_val,
        "tail": m.group("tail"),
        "marked_deleted": False,
    }


def compose_line_1251(base, counter, field, low, high):
    parts = [
        fmt_fixed(base["table"], PREFIX_WIDTHS["table"]),
        base["sp40"],
        fmt_fixed(base["mandt"], PREFIX_WIDTHS["mandt"]),
        fmt_fixed(base["role"], PREFIX_WIDTHS["role"]),
        fmt_fixed(str(counter).rjust(PREFIX_WIDTHS["counter"], "0"), PREFIX_WIDTHS["counter"]),
        fmt_fixed(base["object"], W1251["object"]),
        fmt_fixed(base["auth"], W1251["auth"]),
        fmt_fixed(base.get("variant_pad", ""), W1251["variant_pad"]),
        fmt_fixed(field, W1251["field"]),
        fmt_fixed(low, W1251["low"]),
        fmt_fixed(high, W1251["high"]),
        base.get("modified", " "),
        base.get("deleted", " "),
        base.get("copied", " "),
        base.get("neu", " "),
        fmt_fixed(base.get("node", ""), W1251["node"]),
        base.get("tail", ""),
    ]
    return "".join(parts)


def compose_line_1252(base, counter, varbl, org_value, high):
    return "".join(
        [
            fmt_fixed("AGR_1252", PREFIX_WIDTHS["table"]),
            " " * PREFIX_WIDTHS["sp40"],
            fmt_fixed(base["mandt"], PREFIX_WIDTHS["mandt"]),
            fmt_fixed(base["role"], PREFIX_WIDTHS["role"]),
            fmt_fixed(str(counter).rjust(PREFIX_WIDTHS["counter"], "0"), PREFIX_WIDTHS["counter"]),
            fmt_fixed(varbl, W1252["varbl"]),
            " " * W1252["sp30"],
            fmt_fixed(org_value, W1252["low"]),
            fmt_fixed(high, W1252["high"]),
            base.get("tail", ""),
        ]
    )


# ---------------- rules ----------------

def _xlsx_cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, "g")
    return str(value).strip()


def parse_rules(path, return_meta=False):
    """Parse and validate RULES.xlsx into normalized rule dictionaries."""
    path_obj = Path(path)
    workbook = None
    required_headers = ["action", "table", "mandt", "agr_name", "object", "auth", "field", "low", "high"]
    rules = []
    coverage_rows = []
    roles_touched = set()
    tables_touched = set()
    validation_issues = []
    warning_issues = []
    valid_tables = {"AGR_1251", "AGR_1252"}
    rx_mandt = re.compile(r"^\d{3}$")
    rx_varbl = re.compile(r"^\$[A-Z0-9_]{1,9}$")
    sheet_name = ""
    rows_scanned = 1

    def issue(code, severity, row, table, role, field, *, msg_id="", msg_params=None, detail=None):
        params = msg_params or {}
        rendered = detail if detail is not None else (t(msg_id, **params) if msg_id else "")
        return {
            "code": code,
            "severity": severity,
            "row": row,
            "table": table,
            "role": role,
            "field": field,
            "detail": rendered,
            "msg_id": msg_id,
            "msg_params": params,
        }

    def split_tokens(raw):
        txt_val = raw or ""
        if "|" in txt_val:
            parts = txt_val.split("|")
        elif "," in txt_val:
            parts = txt_val.split(",")
        else:
            parts = [txt_val]
        return [p.strip() for p in parts]

    def make_coverage_row(row_no, raw_action, raw_table, raw_mandt, raw_role, raw_object, raw_auth, raw_field):
        return {
            "row": row_no,
            "action": raw_action,
            "table": raw_table,
            "mandt": raw_mandt,
            "role": raw_role,
            "field": raw_field,
            "object": raw_object,
            "auth": raw_auth,
            "status": "PENDING",
            "matched": 0,
            "deleted": 0,
            "added": 0,
            "replaced": 0,
            "reason_code": "",
            "reason_msg_id": "",
            "reason_params": {},
        }

    if path_obj.suffix.lower() != ".xlsx":
        validation_issues.append(
            issue(
                "VAL-RULES-FILETYPE",
                "SEV2",
                1,
                "",
                "",
                "",
                msg_id="val.rules_filetype",
                msg_params={"name": path_obj.name},
            )
        )
    else:
        try:
            workbook = load_workbook(path, data_only=True, read_only=True)
        except Exception as ex:  # noqa: BLE001
            validation_issues.append(
                issue(
                    "VAL-RULES-OPEN",
                    "SEV2",
                    1,
                    "",
                    "",
                    "",
                    msg_id="val.rules_open_failed",
                    msg_params={"name": path_obj.name, "error": str(ex)},
                )
            )
        else:
            if not workbook.sheetnames:
                validation_issues.append(
                    issue(
                        "VAL-001",
                        "SEV2",
                        1,
                        "",
                        "",
                        "",
                        msg_id="val.file_empty",
                    )
                )
            else:
                sheet_name = "RULES" if "RULES" in workbook.sheetnames else workbook.sheetnames[0]
                sheet = workbook[sheet_name]
                rows_iter = sheet.iter_rows(values_only=True)
                header_row = next(rows_iter, None)
                if header_row is None:
                    validation_issues.append(
                        issue(
                            "VAL-001",
                            "SEV2",
                            1,
                            "",
                            "",
                            "",
                            msg_id="val.file_empty",
                        )
                    )
                else:
                    header_values = [
                        _xlsx_cell_text(value).lstrip("\ufeff").strip().lower()
                        for value in header_row
                    ]
                    header_map = {
                        header: idx
                        for idx, header in enumerate(header_values)
                        if header
                    }
                    missing_headers = sorted(set(required_headers) - set(header_map))
                    if missing_headers:
                        validation_issues.append(
                            issue(
                                "VAL-MISSING-HEADERS",
                                "SEV2",
                                1,
                                "",
                                "",
                                "",
                                msg_id="val.missing_headers",
                                msg_params={"missing": ", ".join(missing_headers)},
                            )
                        )
                    else:
                        for row_no, row_values in enumerate(rows_iter, start=2):
                            rows_scanned = row_no
                            values = list(row_values or [])
                            norm = {
                                key: _xlsx_cell_text(values[idx]) if idx < len(values) else ""
                                for key, idx in header_map.items()
                            }
                            if all((norm.get(key, "") == "") for key in required_headers):
                                continue

                            raw_action = norm.get("action", "")
                            raw_table = norm.get("table", "")
                            raw_mandt = norm.get("mandt", "")
                            raw_role = norm.get("agr_name", "")
                            raw_object = norm.get("object", "")
                            raw_auth = norm.get("auth", "")
                            raw_field = norm.get("field", "")
                            raw_low = norm.get("low", "")
                            raw_high = norm.get("high", "")

                            action = raw_action.lower()
                            table = raw_table.upper()

                            if raw_role:
                                roles_touched.add(raw_role)
                            if raw_table:
                                tables_touched.add(table)

                            row_errors = []
                            row_warnings = []
                            coverage_row = make_coverage_row(
                                row_no,
                                raw_action,
                                raw_table,
                                raw_mandt,
                                raw_role,
                                raw_object,
                                raw_auth,
                                raw_field,
                            )

                            def add_err(
                                code,
                                msg_id,
                                *,
                                row_value=row_no,
                                errors_list=row_errors,
                                table_value=raw_table,
                                role_value=raw_role,
                                field_value=raw_field,
                                **params,
                            ):
                                errors_list.append(
                                    issue(
                                        code,
                                        "SEV2",
                                        row_value,
                                        table_value,
                                        role_value,
                                        field_value,
                                        msg_id=msg_id,
                                        msg_params={"row": row_value, **params},
                                    )
                                )

                            if raw_action == "":
                                add_err("VAL-MISSING-ACTION", "val.row_missing_action")
                            elif action != "replace_list":
                                add_err("VAL-INVALID-ACTION", "val.row_invalid_action", value=raw_action)

                            if raw_table == "":
                                add_err("VAL-MISSING-TABLE", "val.row_missing_table")
                            elif table not in valid_tables:
                                add_err("VAL-INVALID-TABLE", "val.row_invalid_table", value=raw_table)

                            if raw_mandt == "":
                                add_err("VAL-MISSING-MANDT", "val.row_missing_mandt")
                            elif not rx_mandt.fullmatch(raw_mandt):
                                add_err("VAL-MANDT-FORMAT", "val.row_mandt_format", value=raw_mandt)

                            if raw_role == "":
                                add_err("VAL-MISSING-AGR_NAME", "val.row_missing_agr_name")
                            else:
                                if len(raw_role) > 30:
                                    add_err(
                                        "VAL-AGRNAME-LEN",
                                        "val.row_agrname_len",
                                        length=len(raw_role),
                                        value=raw_role,
                                    )
                                if any(ch.isspace() for ch in raw_role):
                                    add_err("VAL-AGRNAME-FORMAT", "val.row_agrname_format", value=raw_role)

                            if raw_field == "":
                                add_err("VAL-MISSING-FIELD", "val.row_missing_field")

                            for col_name, raw_val in (("LOW", raw_low), ("HIGH", raw_high)):
                                for token in split_tokens(raw_val):
                                    if token and len(token) > 40:
                                        add_err(
                                            "VAL-LOWHIGH-LEN",
                                            "val.row_lowhigh_len",
                                            column=col_name,
                                            token=token,
                                            length=len(token),
                                        )

                            if table == "AGR_1251":
                                if raw_object == "":
                                    add_err("VAL-MISSING-OBJECT", "val.row_missing_object")
                                if raw_auth == "":
                                    add_err("VAL-MISSING-AUTH", "val.row_missing_auth")

                                for col_name, raw_val in (
                                    ("OBJECT", raw_object),
                                    ("AUTH", raw_auth),
                                    ("FIELD", raw_field),
                                ):
                                    if raw_val and any(ch.isspace() for ch in raw_val):
                                        add_err(
                                            "VAL-1251-KEY-FORMAT",
                                            "val.row_1251_key_format",
                                            column=col_name,
                                            value=raw_val,
                                        )

                                if raw_object and len(raw_object) > 10:
                                    add_err(
                                        "VAL-OBJECT-LEN",
                                        "val.row_object_len",
                                        length=len(raw_object),
                                        value=raw_object,
                                    )
                                if raw_auth and len(raw_auth) > 12:
                                    add_err("VAL-AUTH-LEN", "val.row_auth_len", length=len(raw_auth), value=raw_auth)
                                if raw_field and len(raw_field) > 10:
                                    add_err(
                                        "VAL-FIELD-LEN",
                                        "val.row_field_len",
                                        length=len(raw_field),
                                        value=raw_field,
                                    )
                            elif table == "AGR_1252":
                                if raw_object or raw_auth:
                                    row_warnings.append(
                                        issue(
                                            "WARN-IGNORED-OBJECT-AUTH",
                                            "SEV3",
                                            row_no,
                                            raw_table,
                                            raw_role,
                                            raw_field,
                                            msg_id="warn.row_ignored_object_auth",
                                            msg_params={"row": row_no},
                                        )
                                    )

                                if raw_field:
                                    if len(raw_field) > 10:
                                        add_err(
                                            "VAL-VARBL-LEN",
                                            "val.row_varbl_len",
                                            length=len(raw_field),
                                            value=raw_field,
                                        )
                                    if not rx_varbl.fullmatch(raw_field):
                                        add_err("VAL-1252-VARBL-FORMAT", "val.row_varbl_format", value=raw_field)

                            try:
                                pairs = split_pairs(
                                    raw_low,
                                    raw_high,
                                    {"row": row_no, "table": table, "role": raw_role, "field": raw_field},
                                )
                            except CodedError as ce:
                                row_errors.append(
                                    issue(
                                        ce.code or "VAL-PAIR",
                                        ce.severity or "SEV2",
                                        row_no,
                                        raw_table,
                                        raw_role,
                                        raw_field,
                                        msg_id="val.row_split_pairs",
                                        msg_params={"row": row_no, "message": ce.message, "details": ce.details or ""},
                                    )
                                )
                                pairs = []

                            if row_errors:
                                first_error = row_errors[0]
                                coverage_row["status"] = "SKIPPED_ERROR"
                                coverage_row["reason_code"] = first_error["code"]
                                coverage_row["reason_msg_id"] = first_error.get("msg_id", "")
                                coverage_row["reason_params"] = first_error.get("msg_params", {})
                                validation_issues.extend(row_errors)
                                warning_issues.extend(row_warnings)
                                coverage_rows.append(coverage_row)
                                continue

                            rules.append(
                                {
                                    "row": row_no,
                                    "action": action,
                                    "table": table,
                                    "mandt": raw_mandt,
                                    "role": raw_role,
                                    "object": raw_object,
                                    "auth": raw_auth,
                                    "field": raw_field,
                                    "pairs": pairs,
                                    "coverage_row": coverage_row,
                                }
                            )
                            coverage_rows.append(coverage_row)
                            warning_issues.extend(row_warnings)

    all_issues = validation_issues + warning_issues
    if workbook is not None:
        workbook.close()
    has_validation_errors = any(it.get("severity") in ("SEV1", "SEV2") for it in all_issues)
    meta = {
        "rules_sheet_detected": sheet_name,
        "rules_stats": {
            "rows_total_including_header": rows_scanned,
            "rules_loaded": len(rules),
            "validation_errors": len([it for it in all_issues if it.get("severity") in ("SEV1", "SEV2")]),
            "validation_warnings": len([it for it in all_issues if it.get("severity") == "SEV3"]),
            "roles_unique": len(roles_touched),
            "tables_touched": sorted(tables_touched),
            "required_columns_ok": not validation_issues or not any(
                it["code"] in {"VAL-MISSING-HEADERS", "VAL-RULES-FILETYPE", "VAL-RULES-OPEN", "VAL-001"}
                for it in validation_issues
            ),
            "sheet_name": sheet_name,
        },
        "validation_issues": all_issues,
        "has_validation_errors": has_validation_errors,
        "coverage_rows": coverage_rows,
    }
    if return_meta:
        return rules, meta
    return rules


# ---------------- processing ----------------

def build_entries(lines):
    """Parse target-table lines while preserving original line order."""
    entries = []
    for idx, ln in enumerate(lines):
        e = parse_entry_1251(ln)
        if not e:
            e = parse_entry_1252(ln)
        if e:
            e["index"] = idx
            e["marked_deleted"] = False
        entries.append(e)  # None means non-target line
    return entries


def key_1251_from_entry(entry):
    return (
        entry["mandt"],
        entry["role"],
        entry["object"],
        entry["auth"],
        entry["field"].strip(),
    )


def key_1252_from_entry(entry):
    return (
        entry["mandt"].strip(),
        entry["role"].strip(),
        entry["varbl"].strip(),
    )


def key_1251_from_rule(rule):
    return (
        fmt_fixed(rule["mandt"], PREFIX_WIDTHS["mandt"]),
        fmt_fixed(rule["role"], PREFIX_WIDTHS["role"]),
        fmt_fixed(rule["object"], W1251["object"]),
        fmt_fixed(rule["auth"], W1251["auth"]),
        fmt_fixed(rule["field"], W1251["field"]).strip(),
    )


def key_1252_from_rule(rule):
    return (
        rule["mandt"].strip(),
        rule["role"].strip(),
        rule["field"].strip(),
    )


def build_entry_indexes(entries):
    index_1251 = defaultdict(list)
    index_1252 = defaultdict(list)
    for entry in entries:
        if not entry:
            continue
        if entry["table_type"] == "AGR_1251":
            index_1251[key_1251_from_entry(entry)].append(entry)
        elif entry["table_type"] == "AGR_1252":
            index_1252[key_1252_from_entry(entry)].append(entry)
    return {
        "AGR_1251": index_1251,
        "AGR_1252": index_1252,
        "stats": {
            "agr_1251_keys": len(index_1251),
            "agr_1252_keys": len(index_1252),
        },
    }


def build_counters_state(entries):
    """Track used counters per (table_type, role)."""
    used = defaultdict(set)
    for e in entries:
        if e and "counter" in e:
            try:
                key = (e["table_type"], e["role"])
                used[key].add(int(e["counter"]))
            except ValueError:
                continue
    return used


def next_counter(key, used):
    """Return the smallest available counter for the role/table and mark it used."""
    n = 1
    s = used[key]
    while n in s:
        n += 1
    s.add(n)
    return n


def handle_rule_1251(r, entries, indexes, counters_used, log_rows, counters, warn_emit=None, coverage_row=None):
    required = ["mandt", "role", "object", "auth", "field"]
    if not all(r.get(k) for k in required):
        counters["warns"] += 1
        detail = t("warn.rule_missing_1251", row=r["row"])
        if coverage_row is not None:
            coverage_row["status"] = "SKIPPED_ERROR"
            coverage_row["reason_code"] = "WARN-RULE"
            coverage_row["reason_msg_id"] = "warn.rule_missing_1251"
            coverage_row["reason_params"] = {"row": r["row"]}
        log_rows.append(
            [
                "WARN-RULE",
                detail,
                "",
            ]
        )
        if warn_emit:
            warn_emit(
                "WARN-RULE",
                detail,
                severity="SEV3",
                rule=r,
                msg_id="warn.rule_missing_1251",
                msg_params={"row": r["row"]},
                legacy=True,
            )
        return

    key = key_1251_from_rule(r)
    field = key[4]

    hits = [entry for entry in indexes["AGR_1251"].get(key, []) if not entry.get("marked_deleted")]
    if coverage_row is not None:
        coverage_row["matched"] = len(hits)

    if not hits:
        counters["warns"] += 1
        detail = t("warn.nobase_1251", key=key, field=field)
        if coverage_row is not None:
            coverage_row["status"] = "NO_BASE"
            coverage_row["reason_code"] = "WARN-NOBASE"
            coverage_row["reason_msg_id"] = "warn.nobase_1251"
            coverage_row["reason_params"] = {"key": key, "field": field}
        log_rows.append(["WARN-NOBASE", detail, ""])
        if warn_emit:
            warn_emit(
                "WARN-NOBASE",
                detail,
                severity="SEV3",
                rule=r,
                msg_id="warn.nobase_1251",
                msg_params={"key": key, "field": field},
                legacy=True,
            )
        return

    base = hits[0]
    befores = []
    for h in hits:
        h["marked_deleted"] = True
        try:
            counters_used[("AGR_1251", h["role"])].discard(int(h["counter"]))
        except ValueError:
            pass
        befores.append(h["raw"])
        counters["deletes"] += 1
        if coverage_row is not None:
            coverage_row["deleted"] += 1

    role_key = ("AGR_1251", base["role"])
    afters = []
    for low_val, high_val in r["pairs"]:
        counter_val = next_counter(role_key, counters_used)
        new_line = compose_line_1251(base, counter_val, field, low_val, high_val)
        counters["adds"] += 1
        afters.append(new_line)
        ne = parse_entry_1251(new_line)
        ne["index"] = len(entries)
        ne["marked_deleted"] = False
        entries.append(ne)
        indexes["AGR_1251"][key_1251_from_entry(ne)].append(ne)
        if coverage_row is not None:
            coverage_row["added"] += 1

    counters["replaces"] += 1
    if coverage_row is not None:
        coverage_row["status"] = "APPLIED"
        coverage_row["replaced"] = 1
    append_replace_logs(befores, afters, log_rows)


def handle_rule_1252(r, entries, indexes, counters_used, log_rows, counters, warn_emit=None, coverage_row=None):
    required = ["mandt", "role", "field"]
    if not all(r.get(k) for k in required):
        counters["warns"] += 1
        detail = t("warn.rule_missing_1252", row=r["row"])
        if coverage_row is not None:
            coverage_row["status"] = "SKIPPED_ERROR"
            coverage_row["reason_code"] = "WARN-RULE"
            coverage_row["reason_msg_id"] = "warn.rule_missing_1252"
            coverage_row["reason_params"] = {"row": r["row"]}
        log_rows.append(["WARN-RULE", detail, ""])
        if warn_emit:
            warn_emit(
                "WARN-RULE",
                detail,
                severity="SEV3",
                rule=r,
                msg_id="warn.rule_missing_1252",
                msg_params={"row": r["row"]},
                legacy=True,
            )
        return

    varbl_clean = r["field"].strip()
    key = key_1252_from_rule(r)

    hits = [entry for entry in indexes["AGR_1252"].get(key, []) if not entry.get("marked_deleted")]
    if coverage_row is not None:
        coverage_row["matched"] = len(hits)

    if not hits:
        counters["warns"] += 1
        detail = t("warn.nobase_1252", key=key)
        if coverage_row is not None:
            coverage_row["status"] = "NO_BASE"
            coverage_row["reason_code"] = "WARN-NOBASE"
            coverage_row["reason_msg_id"] = "warn.nobase_1252"
            coverage_row["reason_params"] = {"key": key}
        log_rows.append(["WARN-NOBASE", detail, ""])
        if warn_emit:
            warn_emit(
                "WARN-NOBASE",
                detail,
                severity="SEV3",
                rule=r,
                msg_id="warn.nobase_1252",
                msg_params={"key": key},
                legacy=True,
            )
        return

    base = hits[0]
    befores = []
    for h in hits:
        h["marked_deleted"] = True
        try:
            counters_used[("AGR_1252", h["role"])].discard(int(h["counter"]))
        except ValueError:
            pass
        befores.append(h["raw"])
        counters["deletes"] += 1
        if coverage_row is not None:
            coverage_row["deleted"] += 1

    role_key = ("AGR_1252", base["role"])
    afters = []
    for low_val, high_val in r["pairs"]:
        counter_val = next_counter(role_key, counters_used)
        new_line = compose_line_1252(base, counter_val, fmt_fixed(varbl_clean, W1252["varbl"]), low_val, high_val)
        counters["adds"] += 1
        afters.append(new_line)
        ne = parse_entry_1252(new_line)
        ne["index"] = len(entries)
        ne["marked_deleted"] = False
        entries.append(ne)
        indexes["AGR_1252"][key_1252_from_entry(ne)].append(ne)
        if coverage_row is not None:
            coverage_row["added"] += 1

    counters["replaces"] += 1
    if coverage_row is not None:
        coverage_row["status"] = "APPLIED"
        coverage_row["replaced"] = 1
    append_replace_logs(befores, afters, log_rows)



