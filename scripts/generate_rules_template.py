#!/usr/bin/env python3
"""Generate the RULES.xlsx template distributed with the release pack."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

HEADERS = ["ACTION", "TABLE", "MANDT", "AGR_NAME", "OBJECT", "AUTH", "FIELD", "LOW", "HIGH"]
EXAMPLE_ROWS = [
    [
        "replace_list",
        "AGR_1251",
        "100",
        "Z:FSBP_CRM_ZSALSPRO_EXT_1004",
        "S_RFC",
        "T-BD08132800",
        "RFC_NAME",
        "0*|A*",
        "9*|Z*",
    ],
    ["replace_list", "AGR_1252", "100", "Z:FSBP_CRM_ZSALSPRO_EXT_1004", "", "", "$WERKS", "0*|A*", "9*|Z*"],
]
COLUMN_WIDTHS = {
    "A": 16,
    "B": 14,
    "C": 10,
    "D": 34,
    "E": 14,
    "F": 16,
    "G": 14,
    "H": 24,
    "I": 24,
}


def build_template(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "RULES"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:I3"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    for row in EXAMPLE_ROWS:
        sheet.append(row)

    for column, width in COLUMN_WIDTHS.items():
        sheet.column_dimensions[column].width = width

    action_validation = DataValidation(type="list", formula1='"replace_list"', allow_blank=False)
    table_validation = DataValidation(type="list", formula1='"AGR_1251,AGR_1252"', allow_blank=False)
    sheet.add_data_validation(action_validation)
    sheet.add_data_validation(table_validation)
    action_validation.add("A2:A5000")
    table_validation.add("B2:B5000")

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()


def main() -> None:
    repo_root = Path(__file__).resolve().parents[1]
    build_template(repo_root / "templates" / "RULES_template.xlsx")


if __name__ == "__main__":
    main()
